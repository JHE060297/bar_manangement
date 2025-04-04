from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from django.http import HttpResponse
from django.db.models import Sum, Count, F, Value, DecimalField
from django.db.models.functions import Coalesce
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import csv
import io
from .models import Reporte
from .serializers import ReporteSerializer
from apps.users.permissions import IsAdmin, IsAdminOCajero
from apps.orders.models import Pedido, DetallePedido, Pago
from apps.inventory.models import Producto


class ReporteViewSet(viewsets.ModelViewSet):
    queryset = Reporte.objects.all().order_by("-fecha_generacion")
    serializer_class = ReporteSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["usuario", "sucursal", "fecha_generacion", "formato"]
    search_fields = ["usuario__nombre", "usuario__apellido", "sucursal__nombre_sucursal"]
    ordering_fields = ["fecha_generacion", "fecha_inicio", "fecha_fin"]

    def get_permissions(self):
        permission_classes = [IsAdminOCajero]  # Solo admin y cajero pueden gestionar reportes
        return [permission() for permission in permission_classes]

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user

        # Si es admin, ve todos los reportes
        if user.is_admin():
            return queryset
        # Si es cajero, solo ve los reportes de su sucursal
        elif user.is_cajero():
            return queryset.filter(sucursal=user.id_sucursal)
        # Para cualquier otro rol, no ve nada
        return Reporte.objects.none()

    def create(self, request, *args, **kwargs):
        # Asignar el usuario actual como el generador del reporte
        request.data["usuario"] = request.user.id_usuario

        # Verificar que el usuario tenga permisos para generar reportes de la sucursal especificada
        sucursal_id = request.data.get("sucursal")
        if not request.user.is_admin() and request.user.id_sucursal.id_sucursal != int(sucursal_id):
            return Response(
                {"error": "No tiene permisos para generar reportes de esta sucursal"}, status=status.HTTP_403_FORBIDDEN
            )

        return super().create(request, *args, **kwargs)

    @action(detail=True, methods=["get"])
    def descargar(self, request, pk=None):
        reporte = self.get_object()

        # Verificar permisos
        if not request.user.is_admin() and request.user.id_sucursal != reporte.sucursal:
            return Response(
                {"error": "No tiene permisos para descargar este reporte"}, status=status.HTTP_403_FORBIDDEN
            )

        # Obtener los datos del reporte
        fecha_inicio = reporte.fecha_inicio
        fecha_fin = reporte.fecha_fin
        sucursal = reporte.sucursal

        # Filtrar pedidos según criterios
        pedidos_query = Pedido.objects.filter(
            created_at__date__gte=fecha_inicio, created_at__date__lte=fecha_fin, estado="pagado"
        )

        # Si el reporte es por sucursal específica
        if not request.user.is_admin() or (request.user.is_admin() and sucursal):
            pedidos_query = pedidos_query.filter(id_mesa__id_sucursal=sucursal)

        # Obtener datos para el reporte
        detalles = (
            DetallePedido.objects.filter(id_pedido__in=pedidos_query)
            .select_related("id_producto", "id_pedido")
            .values("id_producto__nombre_producto")
            .annotate(
                cantidad_total=Sum("cantidad"),
                ingreso_total=Sum(F("cantidad") * F("precio_unitario")),
                costo_total=Sum(F("cantidad") * F("id_producto__precio_compra")),
                ganancia=Sum(F("cantidad") * (F("precio_unitario") - F("id_producto__precio_compra"))),
            )
            .order_by("-ingreso_total")
        )

        # Calcular resumen
        resumen = {
            "total_pedidos": pedidos_query.count(),
            "total_ventas": pedidos_query.aggregate(total=Coalesce(Sum("total"), 0))["total"],
            "total_ganancia": detalles.aggregate(total=Coalesce(Sum("ganancia"), 0))["total"],
        }

        # Generar el archivo según el formato
        if reporte.formato == "xlsx":
            return self._generar_excel(detalles, resumen, reporte)
        elif reporte.formato == "csv":
            return self._generar_csv(detalles, resumen, reporte)
        elif reporte.formato == "pdf":
            # PDF requeriría una librería adicional como ReportLab, por simplicidad no lo implementamos
            return Response({"error": "Formato PDF aún no implementado"}, status=status.HTTP_501_NOT_IMPLEMENTED)

    def _generar_excel(self, detalles, resumen, reporte):
        # Crear un libro de trabajo y una hoja
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Reporte {reporte.fecha_inicio} - {reporte.fecha_fin}"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="404040")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Título
        ws.merge_cells("A1:F1")
        ws["A1"] = f"REPORTE DE VENTAS - {reporte.sucursal.nombre_sucursal}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        # Información del reporte
        ws["A2"] = f"Fecha de generación: {reporte.fecha_generacion.strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A3"] = f"Período: {reporte.fecha_inicio} al {reporte.fecha_fin}"
        ws["A4"] = f"Generado por: {reporte.usuario.get_full_name()}"

        # Resumen
        ws.merge_cells("A6:F6")
        ws["A6"] = "RESUMEN"
        ws["A6"].font = Font(bold=True, size=12)
        ws["A6"].alignment = Alignment(horizontal="center")

        ws["A7"] = "Total de pedidos:"
        ws["B7"] = resumen["total_pedidos"]
        ws["A8"] = "Total de ventas:"
        ws["B8"] = resumen["total_ventas"]
        ws["A9"] = "Total ganancia:"
        ws["B9"] = resumen["total_ganancia"]

        # Encabezados de la tabla de productos
        ws.merge_cells("A11:F11")
        ws["A11"] = "DETALLE DE PRODUCTOS VENDIDOS"
        ws["A11"].font = Font(bold=True, size=12)
        ws["A11"].alignment = Alignment(horizontal="center")

        headers = ["Producto", "Cantidad", "Costo Total", "Ingreso Total", "Ganancia", "Margen (%)"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=12, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Datos de productos
        row = 13
        for detalle in detalles:
            margen = 0
            if detalle["ingreso_total"] > 0:
                margen = (detalle["ganancia"] / detalle["ingreso_total"]) * 100

            ws.cell(row=row, column=1, value=detalle["id_producto__nombre_producto"])
            ws.cell(row=row, column=2, value=detalle["cantidad_total"])
            ws.cell(row=row, column=3, value=float(detalle["costo_total"]))
            ws.cell(row=row, column=4, value=float(detalle["ingreso_total"]))
            ws.cell(row=row, column=5, value=float(detalle["ganancia"]))
            ws.cell(row=row, column=6, value=f"{margen:.2f}%")
            row += 1

        # Ajustar anchos de columna
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        # Guardar en buffer y retornar
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f"attachment; filename=reporte_{reporte.id_reporte}.xlsx"

        return response

    def _generar_csv(self, detalles, resumen, reporte):
        buffer = io.StringIO()
        writer = csv.writer(buffer)

        # Encabezado
        writer.writerow([f"REPORTE DE VENTAS - {reporte.sucursal.nombre_sucursal}"])
        writer.writerow([f"Fecha de generación: {reporte.fecha_generacion.strftime('%Y-%m-%d %H:%M:%S')}"])
        writer.writerow([f"Período: {reporte.fecha_inicio} al {reporte.fecha_fin}"])
        writer.writerow([f"Generado por: {reporte.usuario.get_full_name()}"])
        writer.writerow([])

        # Resumen
        writer.writerow(["RESUMEN"])
        writer.writerow(["Total de pedidos:", resumen["total_pedidos"]])
        writer.writerow(["Total de ventas:", resumen["total_ventas"]])
        writer.writerow(["Total ganancia:", resumen["total_ganancia"]])
        writer.writerow([])

        # Detalle de productos
        writer.writerow(["DETALLE DE PRODUCTOS VENDIDOS"])
        writer.writerow(["Producto", "Cantidad", "Costo Total", "Ingreso Total", "Ganancia", "Margen (%)"])

        for detalle in detalles:
            margen = 0
            if detalle["ingreso_total"] > 0:
                margen = (detalle["ganancia"] / detalle["ingreso_total"]) * 100

            writer.writerow(
                [
                    detalle["id_producto__nombre_producto"],
                    detalle["cantidad_total"],
                    float(detalle["costo_total"]),
                    float(detalle["ingreso_total"]),
                    float(detalle["ganancia"]),
                    f"{margen:.2f}%",
                ]
            )

        # Preparar respuesta
        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type="text/csv")
        response["Content-Disposition"] = f"attachment; filename=reporte_{reporte.id_reporte}.csv"

        return response
